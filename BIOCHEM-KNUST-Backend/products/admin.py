from django.contrib import admin
from django.utils.html import format_html
from django.db.models import Sum, Count, Q, F, Min, Max
from django.urls import path
from django.shortcuts import render
from django.contrib import messages
from django.utils import timezone
from datetime import timedelta
from decimal import Decimal
from .models import (
    Product, ProductPayment, ProductColor, ProductSize, ProductColorImage, 
    ProductAudit, ProductDiscount, ProductDiscountUsage,
    ProductVariantStock, ProductVariantStockAudit
)
from utils.media_mixins import make_media_admin_mixin

# Create media admin mixins for models with file fields
ProductColorImageMediaAdminMixin = make_media_admin_mixin(['color_image'])
ProductMediaAdminMixin = make_media_admin_mixin(['product_image'])


@admin.register(ProductColorImage)
class ProductColorImageAdmin(ProductColorImageMediaAdminMixin, admin.ModelAdmin):
    list_display = ('product', 'color', 'image_preview', 'display_order')
    list_filter = ('product', 'color')
    search_fields = ('product__product_name', 'color__name')
    ordering = ('product', 'display_order')
    
    fieldsets = (
        ('Product & Color', {
            'fields': ('product', 'color', 'display_order'),
        }),
        ('Color-Specific Image', {
            'fields': ('color_image', 'color_image_url'),
            'description': 'Upload an image showing THIS product in THIS color, OR provide an external image URL.',
        }),
    )
    
    def image_preview(self, obj):
        """Display thumbnail of color-specific image"""
        image_url = obj.get_image_url()
        if image_url:
            return format_html(
                '<img src="{}" style="width: 60px; height: 60px; object-fit: cover; border-radius: 4px; border: 1px solid #ddd;" />',
                image_url
            )
        return format_html('<span style="color: gray;">No image</span>')
    image_preview.short_description = 'Image'


@admin.register(ProductColor)
class ProductColorAdmin(admin.ModelAdmin):
    list_display = ('name', 'hex_code_display', 'is_active', 'display_order', 'product_count')
    list_editable = ('is_active', 'display_order')
    list_filter = ('is_active',)
    search_fields = ('name', 'hex_code')
    ordering = ('display_order', 'name')
    
    def hex_code_display(self, obj):
        """Display color with hex code preview"""
        if obj.hex_code:
            return format_html(
                '<span style="background-color: {}; padding: 2px 10px; border: 1px solid #ccc; border-radius: 3px; color: {};">{}</span>',
                obj.hex_code,
                '#fff' if obj.hex_code.lower() in ['#000000', '#000'] else '#000',
                obj.hex_code
            )
        return '-'
    hex_code_display.short_description = 'Color'
    
    def product_count(self, obj):
        """Count products using this color"""
        return obj.products.count()
    product_count.short_description = 'Products'


@admin.register(ProductSize)
class ProductSizeAdmin(admin.ModelAdmin):
    list_display = ('name', 'code', 'is_active', 'display_order', 'product_count')
    list_editable = ('is_active', 'display_order')
    list_filter = ('is_active',)
    search_fields = ('name', 'code')
    ordering = ('display_order', 'name')
    
    def product_count(self, obj):
        """Count products using this size"""
        return obj.products.count()
    product_count.short_description = 'Products'


class ProductVariantStockAuditInline(admin.TabularInline):
    """Inline for viewing stock audit logs"""
    model = ProductVariantStockAudit
    extra = 0
    readonly_fields = ('action', 'quantity_change', 'quantity_after', 'payment', 'notes', 'created_at', 'created_by')
    can_delete = False
    max_num = 0  # Don't allow adding new entries
    verbose_name = "Stock Audit Log"
    verbose_name_plural = "Stock Audit History"
    
    def has_add_permission(self, request, obj=None):
        return False


@admin.register(ProductVariantStock)
class ProductVariantStockAdmin(admin.ModelAdmin):
    """Admin for managing variant-level stock"""
    
    list_display = (
        'product_name',
        'variant_display',
        'stock_display',
        'reserved_quantity',
        'available_display',
        'stock_status',
        'sku',
        'last_restocked_at'
    )
    list_filter = (
        'product',
        'color',
        'size',
    )
    search_fields = (
        'product__product_name',
        'color__name',
        'size__name',
        'size__code',
        'sku'
    )
    ordering = ('product', 'color', 'size')
    list_per_page = 50
    
    readonly_fields = ('created_at', 'updated_at', 'available_quantity_display')
    inlines = [ProductVariantStockAuditInline]
    
    fieldsets = (
        ('Product Variant', {
            'fields': ('product', 'color', 'size', 'sku'),
        }),
        ('Stock Levels', {
            'fields': ('stock_quantity', 'reserved_quantity', 'available_quantity_display', 'low_stock_threshold'),
        }),
        ('Restock Information', {
            'fields': ('last_restocked_at', 'restock_quantity'),
            'classes': ('collapse',),
        }),
        ('Metadata', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',),
        }),
    )
    
    actions = ['restock_action', 'clear_reservations']
    
    def product_name(self, obj):
        return obj.product.product_name
    product_name.short_description = 'Product'
    product_name.admin_order_field = 'product__product_name'
    
    def variant_display(self, obj):
        parts = []
        if obj.color:
            color_html = format_html(
                '<span style="display: inline-block; width: 12px; height: 12px; background: {}; border: 1px solid #ccc; border-radius: 2px; vertical-align: middle;"></span> {}',
                obj.color.hex_code or '#ccc',
                obj.color.name
            )
            parts.append(color_html)
        if obj.size:
            parts.append(format_html('<strong>{}</strong>', obj.size.code))
        return format_html(' / '.join([str(p) for p in parts])) if parts else '-'
    variant_display.short_description = 'Variant'
    
    def stock_display(self, obj):
        return format_html(
            '<span style="font-weight: bold; color: {};">{}</span>',
            '#28a745' if obj.stock_quantity > obj.low_stock_threshold else '#dc3545' if obj.stock_quantity == 0 else '#ffc107',
            obj.stock_quantity
        )
    stock_display.short_description = 'Stock'
    stock_display.admin_order_field = 'stock_quantity'
    
    def available_display(self, obj):
        available = obj.available_quantity
        return format_html(
            '<span style="font-weight: bold;">{}</span>',
            available
        )
    available_display.short_description = 'Available'
    
    def available_quantity_display(self, obj):
        return obj.available_quantity
    available_quantity_display.short_description = 'Available Quantity (Stock - Reserved)'
    
    def stock_status(self, obj):
        if obj.stock_quantity == 0:
            return format_html('<span style="color: #dc3545; font-weight: bold;">❌ OUT OF STOCK</span>')
        elif obj.is_low_stock:
            return format_html('<span style="color: #ffc107; font-weight: bold;">⚠️ LOW STOCK</span>')
        else:
            return format_html('<span style="color: #28a745;">✅ In Stock</span>')
    stock_status.short_description = 'Status'
    
    def restock_action(self, request, queryset):
        """Quick restock action"""
        # This would ideally open a form to input quantity
        # For now, just show a message
        messages.info(request, "To restock, edit each variant individually and update the stock_quantity field.")
    restock_action.short_description = "Restock selected variants"
    
    def clear_reservations(self, request, queryset):
        """Clear all reservations (e.g., abandoned carts)"""
        count = queryset.update(reserved_quantity=0)
        messages.success(request, f"Cleared reservations for {count} variant(s)")
    clear_reservations.short_description = "Clear reservations (release held stock)"


@admin.register(ProductVariantStockAudit)
class ProductVariantStockAuditAdmin(admin.ModelAdmin):
    """Admin for viewing stock audit logs"""
    
    list_display = (
        'variant_display',
        'action',
        'quantity_change_display',
        'quantity_after',
        'payment_link',
        'created_at',
        'created_by'
    )
    list_filter = ('action', 'created_at', 'variant_stock__product')
    search_fields = (
        'variant_stock__product__product_name',
        'variant_stock__sku',
        'notes'
    )
    ordering = ('-created_at',)
    readonly_fields = (
        'variant_stock', 'action', 'quantity_change', 'quantity_after',
        'payment', 'notes', 'created_at', 'created_by'
    )
    
    def has_add_permission(self, request):
        return False
    
    def has_change_permission(self, request, obj=None):
        return False
    
    def has_delete_permission(self, request, obj=None):
        return False
    
    def variant_display(self, obj):
        vs = obj.variant_stock
        parts = [vs.product.product_name]
        if vs.color:
            parts.append(vs.color.name)
        if vs.size:
            parts.append(vs.size.code)
        return ' - '.join(parts)
    variant_display.short_description = 'Variant'
    
    def quantity_change_display(self, obj):
        if obj.quantity_change > 0:
            return format_html('<span style="color: #28a745; font-weight: bold;">+{}</span>', obj.quantity_change)
        elif obj.quantity_change < 0:
            return format_html('<span style="color: #dc3545; font-weight: bold;">{}</span>', obj.quantity_change)
        return '0'
    quantity_change_display.short_description = 'Change'
    
    def payment_link(self, obj):
        if obj.payment:
            return format_html(
                '<a href="/admin/products/productpayment/{}/change/">{}</a>',
                obj.payment.payment_id,
                obj.payment.transaction_validation_code
            )
        return '-'
    payment_link.short_description = 'Payment'


class ProductVariantStockInline(admin.TabularInline):
    """Inline for managing variant stocks within Product admin"""
    model = ProductVariantStock
    extra = 0
    fields = ('color', 'size', 'stock_quantity', 'reserved_quantity', 'low_stock_threshold', 'sku')
    readonly_fields = ('reserved_quantity',)
    verbose_name = "Variant Stock"
    verbose_name_plural = "Variant Stock Levels (Stock per color+size combination)"
    
    def get_queryset(self, request):
        return super().get_queryset(request).select_related('color', 'size')


class ProductColorImageInline(admin.TabularInline):
    """Inline admin for managing color-specific images within Product admin"""
    model = ProductColorImage
    extra = 0
    fields = ('color', 'color_image', 'color_image_url', 'display_order')
    verbose_name = "Color-Specific Image"
    verbose_name_plural = "Color-Specific Images (Upload images for each color variant)"


class ProductDiscountInline(admin.TabularInline):
    """Inline admin for managing discounts within Product admin"""
    model = ProductDiscount
    extra = 0
    fields = ('name', 'discount_type', 'discount_percentage', 'fixed_price', 'target_audience', 'is_active', 'priority')
    readonly_fields = ('times_used',)
    verbose_name = "Discount"
    verbose_name_plural = "Product Discounts (Special pricing for executives, appointees, or specific users)"
    
    def get_queryset(self, request):
        return super().get_queryset(request).select_related('product')


@admin.register(ProductDiscount)
class ProductDiscountAdmin(admin.ModelAdmin):
    """Admin for managing product discounts/special pricing"""
    
    list_display = (
        'name',
        'product_link',
        'discount_display',
        'target_audience',
        'is_active',
        'validity_display',
        'usage_display',
        'priority',
        'created_at',
    )
    list_filter = (
        'is_active',
        'discount_type',
        'target_audience',
        'product',
        'created_at',
    )
    search_fields = ('name', 'product__product_name', 'description')
    ordering = ('-created_at',)
    list_per_page = 25
    list_editable = ('is_active', 'priority')
    
    filter_horizontal = ('specific_users',)
    
    fieldsets = (
        ('Basic Information', {
            'fields': ('name', 'description', 'product'),
        }),
        ('Discount Configuration', {
            'fields': ('discount_type', 'discount_percentage', 'fixed_price'),
            'description': 'Choose either percentage discount OR fixed price. For percentage, enter the discount amount (e.g., 15 for 15% off). For fixed price, enter the special price (e.g., 60.00 instead of 70.00).',
        }),
        ('Target Audience', {
            'fields': ('target_audience', 'specific_users'),
            'description': 'Select who qualifies for this discount. Use "Specific Users" to grant discount to individual users.',
        }),
        ('Validity Period', {
            'fields': ('is_active', 'valid_from', 'valid_until'),
            'description': 'Control when this discount is valid',
        }),
        ('Usage Limits', {
            'fields': ('max_uses', 'times_used', 'max_uses_per_user'),
            'description': 'Limit how many times this discount can be used',
        }),
        ('Priority', {
            'fields': ('priority',),
            'description': 'Higher priority discounts take precedence when multiple discounts apply to the same user.',
        }),
    )
    
    readonly_fields = ('times_used', 'created_at', 'updated_at')
    
    def product_link(self, obj):
        """Display product name as a link"""
        return format_html(
            '<a href="/admin/products/product/{}/change/">{}</a>',
            obj.product.product_id,
            obj.product.product_name[:30] + ('...' if len(obj.product.product_name) > 30 else '')
        )
    product_link.short_description = 'Product'
    
    def discount_display(self, obj):
        """Display the discount value with formatting"""
        if obj.discount_type == 'PERCENTAGE':
            original_price = obj.product.price
            discount_amount = original_price * obj.discount_percentage / 100
            final_price = original_price - discount_amount
            return format_html(
                '<span style="color: green; font-weight: bold;">{}% off</span><br>'
                '<small>GHS {} → GHS {}</small>',
                obj.discount_percentage,
                original_price,
                round(final_price, 2)
            )
        else:
            original_price = obj.product.price
            savings = original_price - obj.fixed_price
            return format_html(
                '<span style="color: blue; font-weight: bold;">GHS {}</span><br>'
                '<small>Was: GHS {} (Save: GHS {})</small>',
                obj.fixed_price,
                original_price,
                savings
            )
    discount_display.short_description = 'Discount'
    
    def validity_display(self, obj):
        """Display validity period"""
        now = timezone.now()
        
        if not obj.is_active:
            return format_html('<span style="color: gray;">Inactive</span>')
        
        if obj.valid_from and now < obj.valid_from:
            return format_html(
                '<span style="color: orange;">Starts: {}</span>',
                obj.valid_from.strftime('%Y-%m-%d')
            )
        
        if obj.valid_until:
            if now > obj.valid_until:
                return format_html('<span style="color: red;">Expired</span>')
            else:
                return format_html(
                    '<span style="color: green;">Until: {}</span>',
                    obj.valid_until.strftime('%Y-%m-%d')
                )
        
        return format_html('<span style="color: green;">✓ Active (No expiry)</span>')
    validity_display.short_description = 'Validity'
    
    def usage_display(self, obj):
        """Display usage statistics"""
        if obj.max_uses:
            return format_html(
                '{} / {} used',
                obj.times_used,
                obj.max_uses
            )
        return format_html('{} used', obj.times_used)
    usage_display.short_description = 'Usage'
    
    def save_model(self, request, obj, form, change):
        if not change:
            obj.created_by = request.user
        super().save_model(request, obj, form, change)
    
    def get_queryset(self, request):
        return super().get_queryset(request).select_related('product', 'created_by')


@admin.register(ProductDiscountUsage)
class ProductDiscountUsageAdmin(admin.ModelAdmin):
    """Admin for viewing discount usage history"""
    
    list_display = ('discount', 'user', 'used_at', 'transaction_link')
    list_filter = ('discount__product', 'discount', 'used_at')
    search_fields = ('user__first_name', 'user__last_name', 'user__phone', 'discount__name')
    ordering = ('-used_at',)
    readonly_fields = ('discount', 'user', 'transaction', 'used_at')
    
    def transaction_link(self, obj):
        if obj.transaction:
            return format_html(
                '<a href="/admin/payments/transaction/{}/change/">{}</a>',
                obj.transaction.id,
                obj.transaction.reference[:20] if obj.transaction.reference else 'View'
            )
        return '-'
    transaction_link.short_description = 'Transaction'
    
    def has_add_permission(self, request):
        return False
    
    def has_change_permission(self, request, obj=None):
        return False
    
    def get_queryset(self, request):
        return super().get_queryset(request).select_related('discount', 'user', 'transaction', 'discount__product')


class ProductAdmin(ProductMediaAdminMixin, admin.ModelAdmin):
    list_display = (
        "product_name",
        "type_of_product",
        "price",
        "stock_quantity",
        "variant_stock_summary",
        "status",
        "is_active",
        "get_availability_status",
        "get_eligibility_summary",
        "get_discounts_summary",
        "created_at",
    )
    list_filter = ("type_of_product", "status", "is_active", "target_audience", "restrict_by_programme", "restrict_by_year", "created_at")
    search_fields = ("product_name", "product_id", "description")
    ordering = ("-created_at",)
    list_per_page = 20
    list_editable = ("status", "is_active", "stock_quantity")
    inlines = [ProductColorImageInline, ProductVariantStockInline, ProductDiscountInline]
    
    fieldsets = (
        ('Basic Information', {
            'fields': ('product_name', 'description', 'type_of_product', 'price', ('product_image', 'product_image_url')),
            'description': 'Upload image OR provide Google Drive/Dropbox link'
        }),
        ('Inventory Management', {
            'fields': ('stock_quantity', 'low_stock_threshold', 'status'),
            'description': 'For products with variants, edit stock in the "Variant Stock Levels" section below. The total stock shown here is auto-calculated from variants.'
        }),
        ('Variant Options', {
            'fields': ('has_colors', 'available_colors', 'has_sizes', 'available_sizes'),
            'description': 'Configure color and size options. Available options automatically update based on stock levels.',
            'classes': ('collapse',)
        }),
        ('Purchase Eligibility / Restrictions', {
            'fields': (
                'target_audience',
                ('restrict_by_programme', 'allowed_programmes'),
                ('restrict_by_year', 'allowed_years'),
                'eligibility_message',
            ),
            'description': 'Control who can purchase this product (staff, executives, appointees, specific programmes or year groups)'
        }),
        ('Sales Timing', {
            'fields': ('sales_start_date', 'sales_end_date'),
            'description': 'Set when this product can be sold (useful for event tickets, pre-orders)',
            'classes': ('collapse',)
        }),
        ('Display Settings', {
            'fields': ('is_active',),
            'description': 'Control product visibility on website'
        }),
    )
    
    filter_horizontal = ('available_colors', 'available_sizes')
    
    actions = ['initialize_variant_stocks', 'sync_availability']
    
    def variant_stock_summary(self, obj):
        """Display variant stock summary"""
        if not obj.has_colors and not obj.has_sizes:
            return format_html('<span style="color: gray;">No variants</span>')
        
        total_variants = obj.variant_stocks.count()
        in_stock = obj.variant_stocks.filter(stock_quantity__gt=0).count()
        low_stock = obj.variant_stocks.filter(
            stock_quantity__gt=0,
            stock_quantity__lte=F('low_stock_threshold')
        ).count()
        out_of_stock = obj.variant_stocks.filter(stock_quantity=0).count()
        
        if total_variants == 0:
            return format_html('<span style="color: orange;">⚠️ No variant stocks set up</span>')
        
        parts = []
        if in_stock > 0:
            parts.append(format_html('<span style="color: #28a745;">✅{}</span>', in_stock))
        if low_stock > 0:
            parts.append(format_html('<span style="color: #ffc107;">⚠️{}</span>', low_stock))
        if out_of_stock > 0:
            parts.append(format_html('<span style="color: #dc3545;">❌{}</span>', out_of_stock))
        
        return format_html(' '.join([str(p) for p in parts]))
    variant_stock_summary.short_description = 'Variants'
    
    def get_availability_status(self, obj):
        """Display availability message in admin"""
        return obj.get_availability_message()
    get_availability_status.short_description = 'Availability'
    
    def get_discounts_summary(self, obj):
        """Display active discounts summary"""
        active_discounts = obj.discounts.filter(is_active=True).count()
        if active_discounts > 0:
            return format_html(
                '<span style="color: green; font-weight: bold;">🏷️ {} active</span>',
                active_discounts
            )
        return format_html('<span style="color: gray;">No discounts</span>')
    get_discounts_summary.short_description = 'Discounts'
    
    def get_eligibility_summary(self, obj):
        """Display eligibility restrictions summary"""
        restrictions = []
        
        if obj.target_audience != 'ALL':
            audience_labels = {
                'STAFF': '👔 Staff',
                'EXECUTIVES': '⭐ Execs',
                'APPOINTEES': '📋 Appointees',
                'EXECUTIVES_AND_APPOINTEES': '⭐📋 Execs+App',
            }
            restrictions.append(audience_labels.get(obj.target_audience, obj.target_audience))
        
        if obj.restrict_by_programme and obj.allowed_programmes:
            restrictions.append(f"🎓 {'/'.join(obj.allowed_programmes)}")
        
        if obj.restrict_by_year and obj.allowed_years:
            years = [f"Y{y}" for y in obj.allowed_years]
            restrictions.append(f"📅 {','.join(years)}")
        
        if restrictions:
            return format_html('<span style="font-size: 0.85em;">{}</span>', ' | '.join(restrictions))
        return format_html('<span style="color: green;">✓ All</span>')
    get_eligibility_summary.short_description = 'Eligibility'
    
    def initialize_variant_stocks(self, request, queryset):
        """Initialize variant stocks for selected products"""
        total_created = 0
        for product in queryset:
            if product.has_colors or product.has_sizes:
                created = product.initialize_variant_stocks(stock_per_variant=0)
                total_created += len(created)
                messages.info(request, f"{product.product_name}: Created {len(created)} variant stock entries")
        
        messages.success(request, f"Total variant stocks created: {total_created}")
    initialize_variant_stocks.short_description = "Initialize variant stocks (create stock entries for all color+size combos)"
    
    def sync_availability(self, request, queryset):
        """Sync available colors/sizes based on stock levels"""
        for product in queryset:
            product.sync_availability_from_stock()
            messages.success(request, f"{product.product_name}: Availability synced from stock levels")
    sync_availability.short_description = "Sync availability from stock (remove out-of-stock colors/sizes)"
    
    def save_model(self, request, obj, form, change):
        """Save model and create audit log"""
        # Track changes
        if change:
            # Get old values
            old_obj = Product.objects.get(pk=obj.pk)
            changed_fields = []
            old_values = {}
            new_values = {}
            
            # Check important fields for changes
            important_fields = ['product_name', 'price', 'stock_quantity', 'status', 'is_active', 'description']
            
            for field in important_fields:
                old_value = getattr(old_obj, field)
                new_value = getattr(obj, field)
                
                if old_value != new_value:
                    changed_fields.append(field)
                    old_values[field] = str(old_value)
                    new_values[field] = str(new_value)
            
            # Determine action type
            action = 'update'
            if 'price' in changed_fields:
                action = 'price_change'
            elif 'stock_quantity' in changed_fields:
                action = 'stock_change'
            elif 'status' in changed_fields:
                action = 'status_change'
            
            # Auto-update status based on stock
            if obj.stock_quantity == 0 and obj.status == 'AVAILABLE':
                obj.status = 'OUT_OF_STOCK'
            elif obj.stock_quantity > 0 and obj.status == 'OUT_OF_STOCK':
                obj.status = 'AVAILABLE'
            
            # Save the product
            super().save_model(request, obj, form, change)
            
            # Create audit log if there were changes
            if changed_fields:
                ProductAudit.objects.create(
                    product=obj,
                    action=action,
                    changed_by=request.user,
                    old_values=old_values,
                    new_values=new_values,
                    changed_fields=changed_fields,
                    ip_address=self._get_client_ip(request),
                    user_agent=request.META.get('HTTP_USER_AGENT', '')[:500]
                )
                
                # Add success message with change details
                if action == 'price_change':
                    old_price = old_values.get('price', 'N/A')
                    new_price = new_values.get('price', 'N/A')
                    messages.success(
                        request,
                        f'Price changed from GH₵ {old_price} to GH₵ {new_price}. Change has been logged.'
                    )
        else:
            # New product creation
            super().save_model(request, obj, form, change)
            
            # Create audit log for creation
            ProductAudit.objects.create(
                product=obj,
                action='create',
                changed_by=request.user,
                new_values={
                    'product_name': obj.product_name,
                    'price': str(obj.price),
                    'stock_quantity': str(obj.stock_quantity),
                    'status': obj.status
                },
                ip_address=self._get_client_ip(request),
                user_agent=request.META.get('HTTP_USER_AGENT', '')[:500]
            )
    
    def _get_client_ip(self, request):
        """Get client IP address from request"""
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')
        return ip
    
    def save_related(self, request, form, formsets, change):
        """
        Save all related objects (inlines) and then sync product availability.
        
        This is called after save_model and after all inline formsets are saved.
        We use this to sync available_colors and available_sizes based on
        the updated variant stock levels.
        """
        super().save_related(request, form, formsets, change)
        
        # Get the product instance
        product = form.instance
        
        # Only sync if product has variants
        if product.has_colors or product.has_sizes:
            # Sync availability from variant stocks
            product.sync_availability_from_stock()
            
            # Show message about availability sync
            total_stock = product.variant_stocks.aggregate(
                total=Sum('stock_quantity')
            )['total'] or 0
            
            in_stock_variants = product.variant_stocks.filter(stock_quantity__gt=0).count()
            total_variants = product.variant_stocks.count()
            
            if total_variants > 0:
                messages.info(
                    request,
                    f'Variant stock updated: {in_stock_variants}/{total_variants} variants in stock '
                    f'(Total: {total_stock} units). Available colors/sizes synced automatically.'
                )


admin.site.register(Product, ProductAdmin)


@admin.register(ProductAudit)
class ProductAuditAdmin(admin.ModelAdmin):
    """Admin interface for viewing product change history"""
    list_display = ('product', 'action', 'changed_by', 'changed_at', 'change_summary')
    list_filter = ('action', 'changed_at', 'changed_by')
    search_fields = ('product__product_name', 'changed_by__username', 'changed_by__first_name', 'changed_by__last_name', 'notes')
    readonly_fields = ('audit_id', 'product', 'action', 'changed_at', 'changed_by', 'old_values', 'new_values', 'changed_fields', 'ip_address', 'user_agent', 'notes')
    ordering = ('-changed_at',)
    list_per_page = 50
    
    fieldsets = (
        ('Change Information', {
            'fields': ('product', 'action', 'changed_at', 'changed_by')
        }),
        ('Changed Fields', {
            'fields': ('changed_fields', 'old_values', 'new_values'),
            'description': 'Details of what was changed'
        }),
        ('Request Metadata', {
            'fields': ('ip_address', 'user_agent'),
            'classes': ('collapse',)
        }),
        ('Notes', {
            'fields': ('notes',),
            'description': 'Additional context about this change'
        }),
    )
    
    def has_add_permission(self, request):
        """Prevent manual creation of audit logs"""
        return False
    
    def has_delete_permission(self, request, obj=None):
        """Prevent deletion of audit logs"""
        return False
    
    def change_summary(self, obj):
        """Display a summary of what changed"""
        if obj.action == 'create':
            # Show initial price for product creation
            if 'price' in obj.new_values:
                return format_html(
                    '<span style="color: #28a745; font-weight: bold;">Created at GH₵ {}</span>',
                    obj.new_values['price']
                )
            return format_html('<span style="color: #28a745;">Product Created</span>')
        
        elif obj.action == 'price_change':
            # Show price change with difference
            if 'price' in obj.old_values and 'price' in obj.new_values:
                old_price = Decimal(obj.old_values['price'])
                new_price = Decimal(obj.new_values['price'])
                diff = new_price - old_price
                percent = (diff / old_price * 100) if old_price > 0 else 0
                
                color = '#28a745' if diff > 0 else '#dc3545'
                symbol = '↑' if diff > 0 else '↓'
                
                # Format the numeric values before passing to format_html
                diff_formatted = f"{diff:+.2f}"
                percent_formatted = f"{percent:+.1f}"
                
                return format_html(
                    '<span style="color: {}; font-weight: bold;">'
                    'GH₵ {} {} GH₵ {} ({}, {}%)'
                    '</span>',
                    color,
                    obj.old_values['price'],
                    symbol,
                    obj.new_values['price'],
                    diff_formatted,
                    percent_formatted
                )
            return format_html('<span style="color: #dc3545;">Price Changed</span>')
        
        elif obj.action == 'stock_change':
            # Show stock change
            if 'stock_quantity' in obj.old_values and 'stock_quantity' in obj.new_values:
                return format_html(
                    '<span style="color: #ffc107;">Stock: {} → {}</span>',
                    obj.old_values['stock_quantity'],
                    obj.new_values['stock_quantity']
                )
            return format_html('<span style="color: #ffc107;">Stock Changed</span>')
        
        elif obj.action == 'status_change':
            # Show status change
            if 'status' in obj.old_values and 'status' in obj.new_values:
                return format_html(
                    '<span style="color: #17a2b8;">Status: {} → {}</span>',
                    obj.old_values['status'],
                    obj.new_values['status']
                )
            return format_html('<span style="color: #17a2b8;">Status Changed</span>')
        
        elif obj.action == 'delete':
            return format_html('<span style="color: #6c757d;">Product Deleted</span>')
        
        else:
            # Generic update
            changed_count = len(obj.changed_fields) if obj.changed_fields else 0
            return format_html(
                '<span style="color: #6c757d;">{} field{} changed</span>',
                changed_count,
                's' if changed_count != 1 else ''
            )
    
    change_summary.short_description = 'Change Summary'


class ProductPaymentAdmin(admin.ModelAdmin):
    """
    Admin interface for ProductPayment records.
    
    SECURITY RESTRICTIONS:
    - NO ADD permission: Payments are created only by the payment system
    - NO DELETE permission: Financial records must be preserved for audit
    - Limited EDIT: Only collection status can be updated by staff
    
    Admins can:
    ✓ View all payment records
    ✓ Search and filter payments
    ✓ Mark merchandise as collected
    ✓ View analytics and reports
    
    Admins CANNOT:
    ✗ Manually create payment records
    ✗ Delete payment records
    ✗ Modify payment amounts, references, or transaction data
    ✗ Change payment status or customer information
    """
    
    list_display = (
        "payment_id",
        "customer_name",
        "phone",
        "reference",
        "transaction_validation_code",
        "product",
        "amount_display",
        "quantity",
        "is_gift_display",
        "academic_year",
        "payed_at",
        "payment_successful",
        "notification_status",
        "merchandise_status",
    )
    list_filter = (
        "payment_successful",
        "is_gift_purchase",
        "merchandise_taken",
        "academic_year",
        "notification_sent",
        "sms_sent",
        "push_sent",
        "payed_at",
        "transaction",
    )
    search_fields = (
        "transaction_validation_code",
        "payment_id",
        "transaction",
        "reference",
        "phone",
        "customer_name",
        "customer_email",
    )
    ordering = ("-payed_at",)
    list_per_page = 20
    
    # ALL fields readonly except collection tracking fields
    readonly_fields = (
        "payment_id",
        "reference",
        "transaction",
        "transaction_record",  # NEW: Link to Transaction
        "transaction_validation_code",
        "payed_at",
        "academic_year",  # Auto-calculated academic year
        "payment_successful",
        "buyer",  # NEW: Who paid
        "product",
        "amount",
        "price_at_purchase",  # Price at time of purchase
        "quantity",
        "customer_name",
        "customer_email",
        "phone",
        "is_gift_purchase",  # NEW: Gift purchase flag
        "purchased_for_phone",  # NEW: Recipient phone
        "purchased_for_user",  # NEW: Recipient user (auto-linked)
        "gift_message",  # NEW: Gift message
        "gift_linked_at",  # NEW: When gift was linked
        "get_total_display",
        "get_variant_selections_display",  # NEW: Display variant preferences
        "preference_change_count",  # NEW: Preference change tracking
        "preferences_last_updated_at",  # NEW: When preferences were last updated
        "preferences_locked",  # NEW: If preferences are locked
        "notification_sent",
        "sms_sent",
        "push_sent",
        "sms_campaign_id",
        "notification_sent_at",
        "notification_error",
    )
    
    fieldsets = (
        ('Payment Information (Read-Only)', {
            'fields': ('payment_id', 'reference', 'transaction', 'transaction_record', 'transaction_validation_code', 'academic_year'),
            'description': '⚠️ These fields are automatically generated by the payment system and cannot be modified.'
        }),
        ('Buyer & Recipient Information (Read-Only)', {
            'fields': ('buyer', 'customer_name', 'customer_email', 'phone', 'is_gift_purchase', 'purchased_for_phone', 'purchased_for_user', 'gift_linked_at', 'gift_message'),
            'description': '🎁 Shows who paid (buyer) and who received (recipient) the product. Gift purchases auto-link when recipient registers.',
            'classes': ('wide',)
        }),
        ('Product & Amount (Read-Only)', {
            'fields': ('product', 'quantity', 'amount', 'price_at_purchase', 'get_total_display', 'get_variant_selections_display'),
            'description': '⚠️ Product and pricing information is set during payment and cannot be changed.'
        }),
        ('Preference Change Tracking (Read-Only)', {
            'fields': ('preference_change_count', 'preferences_last_updated_at', 'preferences_locked'),
            'description': 'ℹ️ Tracks how many times customer has changed their preferences (max 1 change allowed)',
            'classes': ('collapse',)
        }),
        ('Payment Status (Read-Only)', {
            'fields': ('payment_successful', 'payed_at'),
            'description': '⚠️ Payment status is determined by the payment gateway webhook.'
        }),
        ('Notification Status (Read-Only)', {
            'fields': ('notification_sent', 'sms_sent', 'push_sent', 'sms_campaign_id', 'notification_sent_at', 'notification_error'),
            'description': 'ℹ️ Notification delivery status for validation code.',
            'classes': ('collapse',)
        }),
        ('Collection Tracking (Editable)', {
            'fields': ('merchandise_taken', 'merchandise_taken_at', 'taken_by'),
            'description': '✓ Mark merchandise as collected and track who verified the collection.',
            'classes': ('wide',)
        }),
    )
    
    # Disable add permission - payments must come from payment system
    def has_add_permission(self, request):
        """
        Prevent manual creation of payment records.
        Payments must be created through the payment gateway integration.
        """
        return False
    
    # Disable delete permission - preserve financial records
    def has_delete_permission(self, request, obj=None):
        """
        Prevent deletion of payment records.
        Financial records must be preserved for audit and compliance.
        """
        return False
    
    def get_readonly_fields(self, request, obj=None):
        """
        Make all fields readonly except collection tracking fields.
        Only merchandise_taken, merchandise_taken_at, and taken_by can be edited.
        """
        if obj:  # Editing existing record
            # All fields are readonly except collection fields
            return self.readonly_fields
        return self.readonly_fields
    
    def save_model(self, request, obj, form, change):
        """
        Auto-fill 'taken_by' field when marking as collected.
        """
        if change and obj.merchandise_taken and not obj.taken_by:
            obj.taken_by = request.user.get_full_name() or request.user.username
            if not obj.merchandise_taken_at:
                from django.utils import timezone
                obj.merchandise_taken_at = timezone.now()
        super().save_model(request, obj, form, change)
    
    def amount_display(self, obj):
        """Display amount with currency"""
        if obj.amount:
            return f"GH₵ {obj.amount:,.2f}"
        elif obj.product:
            total = obj.product.price * obj.quantity
            return f"GH₵ {total:,.2f}"
        return "N/A"
    amount_display.short_description = 'Amount'
    
    def is_gift_display(self, obj):
        """Display gift purchase status with recipient info"""
        if not obj.is_gift_purchase:
            return format_html('<span style="color: gray;">-</span>')
        
        if obj.purchased_for_user:
            # Gift linked to user
            user_name = obj.purchased_for_user.get_full_name() or obj.purchased_for_user.email
            return format_html(
                '<span style="color: #10B981; font-weight: bold;">🎁 Gift</span><br>'
                '<small style="color: #059669;">✓ Linked to {}</small>',
                user_name
            )
        else:
            # Gift not yet linked
            return format_html(
                '<span style="color: #F59E0B; font-weight: bold;">🎁 Gift</span><br>'
                '<small style="color: #D97706;">⏳ Pending link to {}</small>',
                obj.purchased_for_phone
            )
    is_gift_display.short_description = 'Gift Status'
    
    def price_at_purchase_display(self, obj):
        """Display price at purchase time"""
        if obj.price_at_purchase:
            current_price = obj.product.price if obj.product else None
            price_html = f"GH₵ {obj.price_at_purchase:,.2f}"
            
            # Show if price has changed since purchase
            if current_price and current_price != obj.price_at_purchase:
                diff = current_price - obj.price_at_purchase
                if diff > 0:
                    price_html += f'<br><small style="color: orange;">▲ Price increased</small>'
                else:
                    price_html += f'<br><small style="color: green;">▼ Price decreased</small>'
            
            return format_html(price_html)
        return "N/A"
    price_at_purchase_display.short_description = 'Price at Purchase'
    
    def notification_status(self, obj):
        """Display notification delivery status with icons"""
        if not obj.notification_sent:
            return format_html(
                '<span style="color: gray;">⚠️ Not Sent</span>'
            )
        
        status_parts = []
        
        if obj.push_sent:
            status_parts.append('<span style="color: #10B981; font-weight: bold;">📱 Push</span>')
        
        if obj.sms_sent:
            sms_info = '📧 SMS'
            if obj.sms_campaign_id:
                sms_info += f' ({obj.sms_campaign_id[:8]}...)'
            status_parts.append(f'<span style="color: #0EA5E9; font-weight: bold;">{sms_info}</span>')
        
        if status_parts:
            result = ' + '.join(status_parts)
            if obj.notification_sent_at:
                time_str = obj.notification_sent_at.strftime("%b %d, %I:%M %p")
                result += f'<br><small style="color: gray;">{time_str}</small>'
            return format_html(result)
        
        return format_html('<span style="color: orange;">⚠️ Partial</span>')
    
    notification_status.short_description = 'Notifications'
    
    def merchandise_status(self, obj):
        """Display merchandise collection status with color"""
        if obj.merchandise_taken:
            return format_html(
                '<span style="color: green; font-weight: bold;">✓ Collected</span><br>'
                '<small style="color: gray;">{}</small>',
                obj.merchandise_taken_at.strftime('%b %d, %Y %I:%M %p') if obj.merchandise_taken_at else 'N/A'
            )
        elif obj.payment_successful:
            return format_html(
                '<span style="color: orange; font-weight: bold;">⏳ Awaiting Collection</span>'
            )
        else:
            return format_html(
                '<span style="color: red;">✗ Payment Failed</span>'
            )
    merchandise_status.short_description = 'Merchandise Status'
    
    def get_total_display(self, obj):
        """Display calculated total"""
        total = obj.get_total_amount()
        return f"GH₵ {total:,.2f}"
    get_total_display.short_description = 'Total Amount'
    
    def get_variant_selections_display(self, obj):
        """
        Display variant selections (color/size preferences) in a readable format
        NEW: Reads from ProductPayment.variant_selections
        Handles both raw format (color_id/size_id) and formatted (color/size objects)
        """
        if not obj.variant_selections:
            return format_html('<span style="color: gray;">No preferences set</span>')
        
        from products.models import ProductColor, ProductSize
        
        items = []
        for selection in obj.variant_selections:
            parts = []
            qty = selection.get('quantity', 1)
            
            # Handle both raw format (color_id) and formatted (color object)
            color_id = selection.get('color_id')
            if not color_id and selection.get('color'):
                color_id = selection['color'].get('id')
            
            size_id = selection.get('size_id')
            if not size_id and selection.get('size'):
                size_id = selection['size'].get('id')
            
            # Get color name
            if color_id:
                try:
                    color = ProductColor.objects.get(id=color_id)
                    parts.append(f'<span style="background-color: {color.hex_code}; padding: 2px 6px; border-radius: 3px; color: white; font-size: 11px;">{color.name}</span>')
                except ProductColor.DoesNotExist:
                    parts.append(f'<span style="color: red;">Unknown Color</span>')
            
            # Get size name
            if size_id:
                try:
                    size = ProductSize.objects.get(id=size_id)
                    parts.append(f'<strong>{size.code}</strong>')
                except ProductSize.DoesNotExist:
                    parts.append(f'<span style="color: red;">Unknown Size</span>')
            
            if parts:
                item_str = ' / '.join(parts)
                if qty > 1:
                    item_str += f' <span style="color: gray;">×{qty}</span>'
                items.append(item_str)
        
        if items:
            return format_html('<br>'.join(items))
        
        return format_html('<span style="color: gray;">No preferences</span>')
    
    get_variant_selections_display.short_description = 'Variant Preferences'
    
    def get_urls(self):
        """Add custom URLs for analytics and validation lookup"""
        urls = super().get_urls()
        custom_urls = [
            path('analytics/', self.admin_site.admin_view(self.analytics_view), name='products_analytics'),
            path('validate/', self.admin_site.admin_view(self.validation_lookup_view), name='products_validate'),
            path('export-product-sales/', self.admin_site.admin_view(self.export_product_sales), name='products_export_product_sales'),
        ]
        return custom_urls + urls
    
    def changelist_view(self, request, extra_context=None):
        """Override changelist to add custom action buttons and security notice"""
        extra_context = extra_context or {}
        extra_context['show_analytics_button'] = True
        extra_context['show_validate_button'] = True
        extra_context['security_notice'] = (
            'Payment records are automatically created by the payment system. '
            'You can only update collection status (mark as collected). '
            'Adding or deleting payment records is not allowed to maintain financial integrity.'
        )
        return super().changelist_view(request, extra_context=extra_context)
    
    def analytics_view(self, request):
        """
        Display comprehensive sales analytics with discount tracking.
        
        Revenue calculations are based on actual amounts paid (price_at_purchase or amount),
        NOT current product prices. This ensures accurate historical analysis even when
        prices change over time.
        """
        from decimal import Decimal
        
        # Get selected academic year from query parameter (empty string means "All Years")
        selected_year = request.GET.get('academic_year', None)
        
        # Get all available academic years
        available_years = ProductPayment.objects.filter(
            payment_successful=True,
            academic_year__isnull=False
        ).exclude(academic_year='').values_list('academic_year', flat=True).distinct().order_by('-academic_year')
        
        # Overall statistics (filtered by academic year if selected)
        total_payments = ProductPayment.objects.filter(payment_successful=True)
        if selected_year:
            total_payments = total_payments.filter(academic_year=selected_year)
        
        # Revenue based on actual paid amounts (already accounts for discounts)
        total_revenue = total_payments.aggregate(
            total=Sum('amount')
        )['total'] or Decimal('0')
        
        total_sales_count = total_payments.count()
        total_items_sold = total_payments.aggregate(
            total=Sum('quantity')
        )['total'] or 0
        
        # Collection statistics
        collected_count = total_payments.filter(merchandise_taken=True).count()
        pending_collection = total_payments.filter(merchandise_taken=False).count()
        collection_rate = (collected_count / total_sales_count * 100) if total_sales_count > 0 else 0
        
        # ========== DISCOUNT ANALYTICS ==========
        # Calculate discount statistics
        discounted_payments = total_payments.filter(discount_applied=True)
        discount_stats = discounted_payments.aggregate(
            total_discounted_sales=Count('payment_id'),
            total_original_value=Sum('original_price'),  # What would have been paid without discount
            total_discount_given=Sum('discount_amount'),  # Total savings given to customers
            total_after_discount=Sum('amount'),  # Actual revenue from discounted sales
        )
        
        # Non-discounted sales
        non_discounted_payments = total_payments.filter(discount_applied=False)
        non_discounted_revenue = non_discounted_payments.aggregate(
            total=Sum('amount')
        )['total'] or Decimal('0')
        
        # Calculate discount metrics
        total_discounted_sales = discount_stats['total_discounted_sales'] or 0
        total_original_value = discount_stats['total_original_value'] or Decimal('0')
        total_discount_given = discount_stats['total_discount_given'] or Decimal('0')
        total_after_discount = discount_stats['total_after_discount'] or Decimal('0')
        
        # If original_price is not set, try to calculate from discount_amount
        if total_original_value == 0 and total_after_discount > 0:
            total_original_value = total_after_discount + total_discount_given
        
        # Discount rate (% of sales that had discounts)
        discount_rate = (total_discounted_sales / total_sales_count * 100) if total_sales_count > 0 else 0
        
        # Average discount percentage given
        avg_discount_percent = ((total_discount_given / total_original_value) * 100) if total_original_value > 0 else 0
        
        # Discount breakdown by type
        discount_by_type = discounted_payments.values('discount_type').annotate(
            count=Count('payment_id'),
            total_saved=Sum('discount_amount'),
            total_revenue=Sum('amount')
        ).order_by('-count')
        
        # Discount breakdown by reason (Executive discount, etc.)
        discount_by_reason = discounted_payments.values('discount_reason').annotate(
            count=Count('payment_id'),
            total_saved=Sum('discount_amount'),
            total_revenue=Sum('amount')
        ).order_by('-count')
        
        # Top discounts used (by ProductDiscount)
        top_discounts = discounted_payments.filter(
            discount__isnull=False
        ).values(
            'discount__name',
            'discount__discount_type',
            'discount__discount_percentage',
            'discount__fixed_price'
        ).annotate(
            times_used=Count('payment_id'),
            total_saved=Sum('discount_amount'),
            total_revenue=Sum('amount')
        ).order_by('-times_used')[:10]
        
        # ========== PRODUCT-WISE SALES WITH DISCOUNT INFO ==========
        product_sales = Product.objects.annotate(
            total_sales=Count(
                'productpayment',
                filter=Q(productpayment__payment_successful=True) & 
                       (Q(productpayment__academic_year=selected_year) if selected_year else Q())
            ),
            total_revenue=Sum(
                'productpayment__amount',
                filter=Q(productpayment__payment_successful=True) & 
                       (Q(productpayment__academic_year=selected_year) if selected_year else Q())
            ),
            items_sold=Sum(
                'productpayment__quantity',
                filter=Q(productpayment__payment_successful=True) & 
                       (Q(productpayment__academic_year=selected_year) if selected_year else Q())
            ),
            pending_collection_count=Count(
                'productpayment',
                filter=Q(
                    productpayment__payment_successful=True,
                    productpayment__merchandise_taken=False
                ) & (Q(productpayment__academic_year=selected_year) if selected_year else Q())
            ),
            # NEW: Discount metrics per product
            discounted_sales_count=Count(
                'productpayment',
                filter=Q(productpayment__payment_successful=True, productpayment__discount_applied=True) &
                       (Q(productpayment__academic_year=selected_year) if selected_year else Q())
            ),
            total_discounts_given=Sum(
                'productpayment__discount_amount',
                filter=Q(productpayment__payment_successful=True, productpayment__discount_applied=True) &
                       (Q(productpayment__academic_year=selected_year) if selected_year else Q())
            )
        ).filter(total_sales__gt=0).order_by('-total_revenue')
        
        # Recent sales (last 30 days within selected academic year)
        thirty_days_ago = timezone.now() - timedelta(days=30)
        recent_sales = total_payments.filter(payed_at__gte=thirty_days_ago)
        recent_revenue = recent_sales.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        recent_count = recent_sales.count()
        recent_discounts = recent_sales.filter(discount_applied=True).aggregate(
            total_saved=Sum('discount_amount')
        )['total_saved'] or Decimal('0')
        
        # Sales by product type (filtered by academic year)
        type_sales_query = total_payments
        type_sales = type_sales_query.values('product__type_of_product').annotate(
            count=Count('payment_id'),
            revenue=Sum('amount'),
            discounted_count=Count('payment_id', filter=Q(discount_applied=True)),
            total_discounts=Sum('discount_amount')
        ).order_by('-revenue')
        
        # Top selling products (by quantity)
        top_products_by_quantity = product_sales.order_by('-items_sold')[:5]
        
        # Top selling products (by revenue)
        top_products_by_revenue = product_sales.order_by('-total_revenue')[:5]
        
        # Monthly trend (last 6 months, filtered by academic year)
        monthly_sales = []
        for i in range(6):
            month_start = timezone.now() - timedelta(days=30 * (5 - i))
            month_end = timezone.now() - timedelta(days=30 * (4 - i))
            month_query = total_payments.filter(
                payed_at__gte=month_start,
                payed_at__lt=month_end
            )
            month_data = month_query.aggregate(
                count=Count('payment_id'),
                revenue=Sum('amount'),
                discounts_given=Sum('discount_amount')
            )
            monthly_sales.append({
                'month': month_start.strftime('%B'),
                'count': month_data['count'] or 0,
                'revenue': month_data['revenue'] or Decimal('0'),
                'discounts': month_data['discounts_given'] or Decimal('0'),
            })
        
        # ========== PRICE HISTORY ANALYSIS ==========
        # Show revenue breakdown by different price points
        price_history = []
        for product in product_sales:
            # Get unique price points for this product with sales data
            price_points_qs = ProductPayment.objects.filter(
                product=product,
                payment_successful=True,
                price_at_purchase__isnull=False
            )
            if selected_year:
                price_points_qs = price_points_qs.filter(academic_year=selected_year)
            
            price_points = price_points_qs.values('price_at_purchase').annotate(
                sales_count=Count('payment_id'),
                items_sold=Sum('quantity'),
                revenue=Sum('amount'),
                first_sale=Min('payed_at'),
                last_sale=Max('payed_at'),
                discounted_count=Count('payment_id', filter=Q(discount_applied=True)),
                total_discounts=Sum('discount_amount')
            ).order_by('first_sale')
            
            if price_points.exists():
                # Get audit logs to see when price changes occurred
                price_changes = ProductAudit.objects.filter(
                    product=product,
                    action__in=['price_change', 'create']
                ).order_by('changed_at')
                
                # Convert QuerySet to list of dicts with proper values
                price_changes_list = []
                for change in price_changes:
                    price_changes_list.append({
                        'changed_at': change.changed_at,
                        'action': change.action,
                        'changed_by': change.changed_by,
                        'old_price': change.old_values.get('price') if change.old_values else None,
                        'new_price': change.new_values.get('price') if change.new_values else None,
                    })
                
                # Calculate total for this product
                price_points_list = list(price_points)
                total_product_items = sum(pp['items_sold'] or 0 for pp in price_points_list)
                total_product_revenue = sum(pp['revenue'] or Decimal('0') for pp in price_points_list)
                
                product_price_data = {
                    'product': product,
                    'current_price': product.price,
                    'price_points': price_points_list,
                    'price_changes': price_changes_list,
                    'has_multiple_prices': len(price_points_list) > 1,
                    'total_items': total_product_items,
                    'total_revenue': total_product_revenue,
                }
                
                # Show products with multiple prices OR significant discount activity
                if product_price_data['has_multiple_prices']:
                    price_history.append(product_price_data)
        
        context = {
            'title': 'Product Sales Analytics',
            'available_years': list(available_years),
            'selected_year': selected_year,
            'total_revenue': total_revenue,
            'total_sales_count': total_sales_count,
            'total_items_sold': total_items_sold,
            'collected_count': collected_count,
            'pending_collection': pending_collection,
            'collection_rate': collection_rate,
            'product_sales': product_sales,
            'recent_revenue': recent_revenue,
            'recent_count': recent_count,
            'recent_discounts': recent_discounts,
            'type_sales': type_sales,
            'top_products_by_quantity': top_products_by_quantity,
            'top_products_by_revenue': top_products_by_revenue,
            'monthly_sales': monthly_sales,
            'price_history': price_history,
            'opts': self.model._meta,
            
            # NEW: Discount analytics context
            'discount_stats': {
                'total_discounted_sales': total_discounted_sales,
                'total_original_value': total_original_value,
                'total_discount_given': total_discount_given,
                'total_after_discount': total_after_discount,
                'non_discounted_revenue': non_discounted_revenue,
                'discount_rate': discount_rate,
                'avg_discount_percent': avg_discount_percent,
            },
            'discount_by_type': discount_by_type,
            'discount_by_reason': discount_by_reason,
            'top_discounts': top_discounts,
        }
        
        return render(request, 'admin/products/analytics.html', context)
    
    def validation_lookup_view(self, request):
        """
        Enhanced validation and collection lookup with per-variant tracking.
        
        Supports:
        - Full collection (all variants at once)
        - Partial collection (specific variants/quantities)
        - Substitution tracking (gave different color/size)
        - Marking items as unavailable
        - Bulk validation (lookup and collect multiple codes at once)
        """
        from products.payment_service import ProductPaymentService
        import json
        
        payment = None
        error = None
        validation_code = None
        collection_status = None
        bulk_results = None
        action_result = None
        
        if request.method == 'POST':
            validation_code = request.POST.get('validation_code', '').strip().upper()
            action = request.POST.get('action', 'search')
            mode = request.POST.get('mode', 'single')
            
            # Handle bulk operations
            if mode == 'bulk' and action in ['bulk_lookup', 'bulk_collect']:
                codes = request.POST.getlist('codes[]')
                # Filter valid codes (6 alphanumeric characters)
                import re
                valid_codes = [c.strip().upper() for c in codes if c.strip() and re.match(r'^[A-Za-z0-9]{6}$', c.strip())]
                
                if not valid_codes:
                    messages.error(request, "Please enter at least one valid validation code")
                else:
                    results = []
                    found_count = 0
                    not_found_count = 0
                    pending_count = 0
                    pending_codes = []
                    collected_in_bulk = 0
                    
                    for code in valid_codes:
                        try:
                            status = ProductPaymentService.get_collection_status(code)
                            
                            if not status.get('payment_successful'):
                                results.append({
                                    'code': code,
                                    'error': 'Payment not successful'
                                })
                                not_found_count += 1
                                continue
                            
                            summary = status.get('collection_summary', {})
                            already_collected = summary.get('overall_status') == 'collected'
                            pending_items = summary.get('total_pending', 0)
                            
                            result_item = {
                                'code': code,
                                'customer_name': status.get('customer_name', 'N/A'),
                                'product_name': status.get('product_name', 'N/A'),
                                'amount': status.get('amount_paid', 0),
                                'total_items': summary.get('total_ordered', 0),
                                'pending_items': pending_items,
                                'already_collected': already_collected
                            }
                            
                            # If bulk collect action, mark as collected
                            if action == 'bulk_collect' and not already_collected:
                                try:
                                    collect_result = ProductPaymentService.mark_merchandise_collected(
                                        validation_code=code,
                                        collected_by=request.user.get_full_name() or request.user.username
                                    )
                                    result_item['collected_now'] = True
                                    result_item['already_collected'] = True
                                    result_item['pending_items'] = 0
                                    collected_in_bulk += 1
                                except ValueError as e:
                                    result_item['collect_error'] = str(e)
                            
                            results.append(result_item)
                            found_count += 1
                            
                            if not already_collected and action != 'bulk_collect':
                                pending_count += 1
                                pending_codes.append(code)
                            
                        except ValueError as e:
                            results.append({
                                'code': code,
                                'error': str(e)
                            })
                            not_found_count += 1
                        except Exception as e:
                            results.append({
                                'code': code,
                                'error': f'Error: {str(e)}'
                            })
                            not_found_count += 1
                    
                    bulk_results = {
                        'total': len(valid_codes),
                        'found': found_count,
                        'not_found': not_found_count,
                        'pending_collection': pending_count if action == 'bulk_lookup' else 0,
                        'pending_codes': pending_codes if action == 'bulk_lookup' else [],
                        'results': results
                    }
                    
                    if action == 'bulk_collect':
                        messages.success(request, f'⚡ Bulk collection complete! {collected_in_bulk} items collected.')
                    else:
                        messages.info(request, f'🔍 Found {found_count} of {len(valid_codes)} codes. {pending_count} pending collection.')
            
            elif validation_code:
                try:
                    # Get current collection status
                    collection_status = ProductPaymentService.get_collection_status(validation_code)
                    
                    if not collection_status.get('payment_successful'):
                        error = "Payment was not successful for this code"
                        messages.error(request, error)
                        collection_status = None
                    
                    elif action == 'collect_all':
                        # Mark all variants as collected
                        try:
                            result = ProductPaymentService.mark_merchandise_collected(
                                validation_code=validation_code,
                                collected_by=request.user.get_full_name() or request.user.username
                            )
                            if result.get('fully_collected'):
                                messages.success(request, f'✅ All items collected successfully! Customer: {result.get("customer_name", "N/A")}')
                            else:
                                messages.success(request, f'✅ Items collected. {result["collection_summary"]["total_collected"]}/{result["collection_summary"]["total_ordered"]} items collected.')
                            
                            # Refresh status
                            collection_status = ProductPaymentService.get_collection_status(validation_code)
                            
                        except ValueError as e:
                            messages.error(request, f'❌ Error: {str(e)}')
                    
                    elif action == 'collect_variant':
                        # Collect specific variant
                        variant_index = int(request.POST.get('variant_index', 0))
                        quantity = int(request.POST.get('quantity', 1))
                        substituted = request.POST.get('substituted', 'false') == 'true'
                        substitution_notes = request.POST.get('substitution_notes', '')
                        
                        try:
                            result = ProductPaymentService.mark_merchandise_collected(
                                validation_code=validation_code,
                                collected_by=request.user.get_full_name() or request.user.username,
                                variant_collections=[{
                                    'variant_index': variant_index,
                                    'quantity': quantity,
                                    'substituted': substituted,
                                    'substitution_notes': substitution_notes
                                }]
                            )
                            summary = result.get('collection_summary', {})
                            messages.success(request, f'✅ Variant collected! ({summary.get("total_collected", 0)}/{summary.get("total_ordered", 0)} items total)')
                            
                            # Refresh status
                            collection_status = ProductPaymentService.get_collection_status(validation_code)
                            
                        except ValueError as e:
                            messages.error(request, f'❌ Error: {str(e)}')
                    
                    elif action == 'mark_unavailable':
                        # Mark variant as unavailable
                        variant_index = int(request.POST.get('variant_index', 0))
                        reason = request.POST.get('reason', '')
                        
                        try:
                            result = ProductPaymentService.mark_variant_unavailable(
                                validation_code=validation_code,
                                variant_index=variant_index,
                                reason=reason,
                                marked_by=request.user.get_full_name() or request.user.username
                            )
                            messages.warning(request, f'⚠️ Variant marked as unavailable.')
                            
                            # Refresh status
                            collection_status = ProductPaymentService.get_collection_status(validation_code)
                            
                        except ValueError as e:
                            messages.error(request, f'❌ Error: {str(e)}')
                    
                    # For backward compatibility, also get ProductPayment for display
                    if collection_status and collection_status.get('source') == 'product_payment':
                        try:
                            payment = ProductPayment.objects.select_related('product').get(
                                transaction_validation_code=validation_code
                            )
                        except ProductPayment.DoesNotExist:
                            pass
                
                except ValueError as e:
                    error = str(e)
                    messages.error(request, error)
                except Exception as e:
                    error = f"An error occurred: {str(e)}"
                    messages.error(request, error)
            else:
                error = "Please enter a validation code"
                messages.error(request, error)
        
        context = {
            'title': 'Validate & Collect Merchandise',
            'payment': payment,
            'collection_status': collection_status,
            'bulk_results': bulk_results,
            'error': error,
            'validation_code': validation_code,
            'opts': self.model._meta,
        }
        
        return render(request, 'admin/products/validate.html', context)
    
    def export_product_sales(self, request):
        """
        Export product sales to Excel with detailed information:
        - Customer Name
        - Color Choice
        - Size Choice
        - Cost Paid
        - Committee/Position (Executive or Appointee)
        
        Filters: product_id (required), academic_year (optional)
        """
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        from products.models import ProductColor, ProductSize
        from executives.models import Executive, Appointee
        
        # Get filter parameters
        product_id = request.GET.get('product_id')
        academic_year = request.GET.get('academic_year', '')
        
        if not product_id:
            messages.error(request, "Product ID is required")
            return self.analytics_view(request)
        
        try:
            product = Product.objects.get(product_id=product_id)
        except Product.DoesNotExist:
            messages.error(request, "Product not found")
            return self.analytics_view(request)
        
        # Get sales data
        sales = ProductPayment.objects.filter(
            product=product,
            payment_successful=True
        ).select_related('product').order_by('-payed_at')
        
        if academic_year:
            sales = sales.filter(academic_year=academic_year)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{product.product_name[:25]} Sales"
        
        # Header styling
        header_fill = PatternFill(start_color="417690", end_color="417690", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Define headers
        headers = [
            'Customer Name',
            'Email',
            'Phone',
            'Committee',
            'Position/Role',
            'Color Choice',
            'Size Choice',
            'Quantity',
            'Price at Purchase (GH₵)',  # NEW: Historical price
            'Cost Paid (GH₵)',
            'Payment Date',
            'Validation Code',
            'Collected',
            'Academic Year'
        ]
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Write data rows
        row_num = 2
        for sale in sales:
            # Extract variant selections
            color_choices = []
            size_choices = []
            
            if sale.variant_selections:
                for selection in sale.variant_selections:
                    if selection.get('color_id'):
                        try:
                            color = ProductColor.objects.get(id=selection['color_id'])
                            qty = selection.get('quantity', 1)
                            color_choices.append(f"{color.name} (×{qty})" if qty > 1 else color.name)
                        except ProductColor.DoesNotExist:
                            color_choices.append("Unknown")
                    
                    if selection.get('size_id'):
                        try:
                            size = ProductSize.objects.get(id=selection['size_id'])
                            qty = selection.get('quantity', 1)
                            size_choices.append(f"{size.code} (×{qty})" if qty > 1 else size.code)
                        except ProductSize.DoesNotExist:
                            size_choices.append("Unknown")
            
            color_str = ", ".join(color_choices) if color_choices else "N/A"
            size_str = ", ".join(size_choices) if size_choices else "N/A"
            
            # Convert phone number to string
            phone_str = str(sale.phone) if sale.phone else "N/A"
            
            # Get committee and position information
            committee_str = "N/A"
            position_str = "N/A"
            
            if sale.customer_email:
                try:
                    from accounts.models import CustomUser
                    from django.db.models import Q
                    
                    # Find user by email (check both student_email and personal_email)
                    user = CustomUser.objects.filter(
                        Q(student_email=sale.customer_email) | Q(personal_email=sale.customer_email)
                    ).first()
                    
                    if user:
                        # Check if user is an Executive
                        executive = Executive.objects.select_related('position').filter(
                            user=user, 
                            is_active=True
                        ).first()
                        
                        if executive:
                            committee_str = "Executive Board"
                            position_str = executive.position.name if executive.position else "N/A"
                        else:
                            # Check if user is an Appointee
                            appointee = Appointee.objects.select_related('committee').filter(
                                user=user, 
                                is_active=True
                            ).first()
                            
                            if appointee:
                                committee_str = appointee.committee.name if appointee.committee else "N/A"
                                position_str = appointee.get_role_display() if appointee.role else "N/A"
                except Exception:
                    pass  # Keep default N/A values
            
            # Write row data
            ws.cell(row=row_num, column=1).value = sale.customer_name or "N/A"
            ws.cell(row=row_num, column=2).value = sale.customer_email or "N/A"
            ws.cell(row=row_num, column=3).value = phone_str
            ws.cell(row=row_num, column=4).value = committee_str
            ws.cell(row=row_num, column=5).value = position_str
            ws.cell(row=row_num, column=6).value = color_str
            ws.cell(row=row_num, column=7).value = size_str
            ws.cell(row=row_num, column=8).value = sale.quantity
            ws.cell(row=row_num, column=9).value = float(sale.price_at_purchase) if sale.price_at_purchase else 0.0  # NEW: Historical price
            ws.cell(row=row_num, column=10).value = float(sale.amount) if sale.amount else 0.0
            ws.cell(row=row_num, column=11).value = sale.payed_at.strftime('%Y-%m-%d %H:%M') if sale.payed_at else "N/A"
            ws.cell(row=row_num, column=12).value = sale.transaction_validation_code or "N/A"
            ws.cell(row=row_num, column=13).value = "Yes" if sale.merchandise_taken else "No"
            ws.cell(row=row_num, column=14).value = sale.academic_year or "N/A"
            
            row_num += 1
        
        # Auto-adjust column widths
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = len(headers[col_num - 1])
            for cell in ws[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Prepare response
        filename = f"{product.product_name.replace(' ', '_')}_Sales"
        if academic_year:
            filename += f"_{academic_year.replace('/', '-')}"
        filename += ".xlsx"
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response


admin.site.register(ProductPayment, ProductPaymentAdmin)
