"""
KNUST Admissions Data Scraper
Fetches and updates admission data from the KNUST admissions portal
"""
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os
import logging
from datetime import datetime
from django.conf import settings
from django.utils import timezone

logger = logging.getLogger('scraper')

# KNUST Admissions URL
URL = "https://apps.knust.edu.gh/admissions/check/Home/Undergraduates"

# Excel file path in media directory
EXCEL_DIR = os.path.join(settings.MEDIA_ROOT, 'admissions')
EXCEL_FILE = os.path.join(EXCEL_DIR, 'knust_admissions.xlsx')


class KNUSTAdmissionsScraper:
    """Scraper for KNUST admissions data"""
    
    def __init__(self):
        self.url = URL
        self.excel_file = EXCEL_FILE
        self.headers = ["Applicant ID", "Name", "Programme", "Academic Year", "Admission Status", "Source", "Fetched At"]
        
        # Ensure directory exists
        os.makedirs(EXCEL_DIR, exist_ok=True)
    
    def fetch_data(self):
        """Fetch and scrape Computer Science applicant data from the KNUST admissions page."""
        try:
            
            response = requests.get(self.url, timeout=20)
            response.raise_for_status()
            
            soup = BeautifulSoup(response.text, "html.parser")
            
            # Find all tables with the class 'datatableApplicants'
            tables = soup.find_all("table", class_="datatableApplicants")
            
            if not tables:
                raise Exception("Could not find admissions tables on the page.")
            
            data_rows = []
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # Map of section IDs to admission group names
            admission_groups = {
                "v-pills-international-applicants": "International Applicants",
                "v-pills-fee-paying-other-applicants": "Fee Paying (Mature/Other)",
                "v-pills-mature-applicants": "Mature/Other/Foreign Results",
                "v-pills-wassce-applicants": "SSSCE/WASSCE",
                "v-pills-fee-paying-wassce-applicants": "Fee Paying SSSCE/WASSCE",
                "v-pills-less-endowed-applicants": "Less Endowed Schools",
                "v-pills-NMTC-applicants": "NMTC Upgrade"
            }
            
            # Process each table
            for table in tables:
                # Find the parent section to determine admission group
                parent_section = table.find_parent("div", class_="tab-pane")
                admission_group = "Unknown"
                
                if parent_section and parent_section.get("id"):
                    section_id = parent_section.get("id")
                    admission_group = admission_groups.get(section_id, "Unknown")
                
                # Process rows in this table
                for row in table.find_all("tr")[1:]:  # Skip header row
                    cols = row.find_all("td")
                    if cols and len(cols) >= 4:  # Need at least 4 columns (excluding action button)
                        # Extract text from columns
                        serial_num = cols[0].text.strip()
                        applicant_id = cols[1].text.strip()
                        name = cols[2].text.strip()
                        programme = cols[3].text.strip()
                        
                        # Filter for Computer Science and Information Technology programs
                        # Be specific to avoid catching Business IT
                        if ("BSC. COMPUTER SCIENCE" in programme.upper() or 
                            "BSC. INFORMATION TECHNOLOGY" in programme.upper() or
                            "B.SC. COMPUTER SCIENCE" in programme.upper() or
                            "B.SC. INFORMATION TECHNOLOGY" in programme.upper()):
                            
                            # Add row: [Applicant ID, Name, Programme, Admission Group, Timestamp]
                            data_rows.append([applicant_id, name, programme, admission_group, timestamp])
          
            return data_rows, None
            
        except requests.RequestException as e:
            error_msg = f"Network error while fetching data: {str(e)}"
          
            return [], error_msg
        except Exception as e:
            error_msg = f"Error parsing data: {str(e)}"
        
            return [], error_msg
    
    def create_excel(self, rows):
        """Create new Excel file with styling."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Admissions Data"
        
        # Add headers with styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        ws.append(self.headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add data rows
        for row in rows:
            ws.append(row)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(self.excel_file)
      
        return len(rows)
    
    def update_excel(self, rows):
        """Insert new rows without duplicating existing students."""
        wb = load_workbook(self.excel_file)
        ws = wb.active
        
        existing_ids = set()
        id_col = 1  # Applicant ID is first column
        
        # Read existing applicant IDs
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                existing_ids.add(str(row[0]).strip())
        
        added = 0
        updated = 0
        
        # Add only rows that don't already exist
        for row in rows:
            applicant_id = str(row[0]).strip()
            if applicant_id not in existing_ids:
                ws.append(row)
                existing_ids.add(applicant_id)
                added += 1
            else:
                # Optionally update existing row
                updated += 1
        
        wb.save(self.excel_file)
 
        return added, updated
    
    def scrape_and_update(self):
        """
        Main method to scrape data, save to database, and update Excel file.
        Returns: (success, message, stats_dict)
        """
        try:
            # Import here to avoid circular import
            from .models import KNUSTAdmission
            
            # Fetch data
            scraped_data, error = self.fetch_data()
            
            if error:
                return False, error, {}
            
            if not scraped_data:
                return False, "No Computer Science or Information Technology data found on the KNUST admissions page.", {}
            
            # Track duplicates and analyze them
            seen_applicants = {}  # {applicant_id: {'name': name, 'programme': programme, 'group': group}}
            duplicates_same_name = []  # True duplicates (same ID, same name)
            duplicates_diff_name = []  # Potential issues (same ID, different name)
            unique_data = []  # Deduplicated data
            
            for row in scraped_data:
                if len(row) >= 4:
                    applicant_id = str(row[0]).strip()
                    name = str(row[1]).strip()
                    programme = row[2]
                    admission_group = row[3]
                    
                    if applicant_id in seen_applicants:
                        existing = seen_applicants[applicant_id]
                        # Check if names match (case-insensitive)
                        if existing['name'].lower() == name.lower():
                            duplicates_same_name.append({
                                'applicant_id': applicant_id,
                                'name': name,
                                'existing_group': existing['group'],
                                'new_group': admission_group
                            })
                        else:
                            duplicates_diff_name.append({
                                'applicant_id': applicant_id,
                                'existing_name': existing['name'],
                                'new_name': name,
                                'existing_group': existing['group'],
                                'new_group': admission_group
                            })
                    else:
                        seen_applicants[applicant_id] = {
                            'name': name,
                            'programme': programme,
                            'group': admission_group
                        }
                        unique_data.append(row)
            
            # Save unique data to database
            db_added = 0
            db_updated = 0
            
            for row in unique_data:
                applicant_id = str(row[0]).strip()
                name = str(row[1]).strip()
                programme = row[2]
                admission_group = row[3]
                
                # Get admission status from group name
                admission_status = KNUSTAdmission.get_admission_status_from_group(admission_group)
                
                # Use get_or_create with academic_year to prevent duplicates
                admission, created = KNUSTAdmission.objects.get_or_create(
                    applicant_id=applicant_id,
                    academic_year=KNUSTAdmission._meta.get_field('academic_year').default(),
                    defaults={
                        'name': name,
                        'programme': programme,
                        'admission_status': admission_status,
                        'status': admission_group,
                        'source': 'SCRAPER',
                    }
                )
                
                if created:
                    db_added += 1
                else:
                    # Update existing record
                    admission.name = name
                    admission.programme = programme
                    admission.admission_status = admission_status
                    admission.status = admission_group
                    admission.save()
                    db_updated += 1
            
            # Create or update Excel file with unique data only
            if not os.path.exists(self.excel_file):
                total = self.create_excel(unique_data)
                stats = {
                    'total_fetched': len(scraped_data),
                    'unique_records': len(unique_data),
                    'new_records': total,
                    'existing_records': 0,
                    'db_added': db_added,
                    'db_updated': db_updated,
                    'duplicates_same_name': len(duplicates_same_name),
                    'duplicates_diff_name': len(duplicates_diff_name),
                    'duplicate_details': {
                        'same_name': duplicates_same_name[:10],  # Limit to first 10
                        'diff_name': duplicates_diff_name[:10]
                    },
                    'file_path': self.excel_file,
                    'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                }
                message = f"✅ Excel file created with {total} records. Database: {db_added} new, {db_updated} updated."
            else:
                added, existing = self.update_excel(unique_data)
                stats = {
                    'total_fetched': len(scraped_data),
                    'unique_records': len(unique_data),
                    'new_records': added,
                    'existing_records': existing,
                    'db_added': db_added,
                    'db_updated': db_updated,
                    'duplicates_same_name': len(duplicates_same_name),
                    'duplicates_diff_name': len(duplicates_diff_name),
                    'duplicate_details': {
                        'same_name': duplicates_same_name[:10],  # Limit to first 10
                        'diff_name': duplicates_diff_name[:10]
                    },
                    'file_path': self.excel_file,
                    'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                }
                message = f"✅ Excel: {added} new, {existing} existing. Database: {db_added} new, {db_updated} updated."
            
            # Add duplicate info to message
            if duplicates_same_name:
                message += f"\n👥 {len(duplicates_same_name)} duplicate(s) with same name (skipped)."
            if duplicates_diff_name:
                message += f"\n⚠️ {len(duplicates_diff_name)} duplicate ID(s) with DIFFERENT names detected!"
            
            return True, message, stats
            
        except Exception as e:
            error_msg = f"Unexpected error: {str(e)}"
          
            import traceback
            traceback.print_exc()
            return False, error_msg, {}
    
    def get_excel_data(self, limit=100):
        """
        Read data from Excel file for display.
        Returns: (headers, rows)
        """
        try:
            if not os.path.exists(self.excel_file):
                return [], []
            
            wb = load_workbook(self.excel_file, read_only=True)
            ws = wb.active
            
            # Get headers
            headers = [cell.value for cell in ws[1]]
            
            # Get data rows (limit to avoid loading too much)
            rows = []
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                if i >= limit:
                    break
                if row and row[0]:  # Skip empty rows
                    rows.append(row)
            
            wb.close()
            return headers, rows
            
        except Exception as e:
      
            return [], []
    
    def clean_excel_duplicates(self):
        """
        Remove duplicate entries from Excel file based on applicant_id.
        Returns: (removed_count, total_after_cleanup)
        """
        if not os.path.exists(self.excel_file):
            return 0, 0
        
        try:
            wb = load_workbook(self.excel_file)
            ws = wb.active
            
            seen_ids = set()
            rows_to_delete = []
            
            # Find duplicate rows (skip header row 1)
            for row_num in range(2, ws.max_row + 1):
                applicant_id = ws.cell(row=row_num, column=1).value
                if applicant_id:
                    applicant_id = str(applicant_id).strip()
                    if applicant_id in seen_ids:
                        rows_to_delete.append(row_num)
                    else:
                        seen_ids.add(applicant_id)
            
            # Delete duplicate rows in reverse order to maintain row indices
            for row_num in reversed(rows_to_delete):
                ws.delete_rows(row_num)
            
            wb.save(self.excel_file)
            wb.close()
            
            logger.info(f"Removed {len(rows_to_delete)} duplicate rows from Excel file")
            return len(rows_to_delete), ws.max_row - 1
            
        except Exception as e:
            logger.error(f"Error cleaning Excel duplicates: {str(e)}")
            return 0, 0
    
    def sync_excel_with_database(self):
        """
        Rebuild Excel file from database to ensure they are in sync.
        This is the authoritative sync - database is the source of truth.
        Returns: (success, message, total_records)
        """
        try:
            from .models import KNUSTAdmission
            
            # Get all records from database
            admissions = KNUSTAdmission.objects.all().order_by('-fetched_at')
            
            if not admissions.exists():
                return False, "No records in database to sync", 0
            
            # Create new Excel file from database
            wb = Workbook()
            ws = wb.active
            ws.title = "Admissions Data"
            
            # Add headers with styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            
            ws.append(self.headers)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Add all database records
            for admission in admissions:
                ws.append([
                    admission.applicant_id,
                    admission.name,
                    admission.programme,
                    admission.academic_year,
                    admission.get_admission_status_display(),
                    admission.get_source_display(),
                    admission.fetched_at.strftime("%Y-%m-%d %H:%M:%S") if admission.fetched_at else ""
                ])
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(self.excel_file)
            wb.close()
            
            total = admissions.count()
            logger.info(f"Synced Excel file with database: {total} records")
            return True, f"✅ Excel file rebuilt from database with {total} records", total
            
        except Exception as e:
            error_msg = f"Error syncing Excel with database: {str(e)}"
            logger.error(error_msg)
            import traceback
            traceback.print_exc()
            return False, error_msg, 0
            
            logger.info(f"Removed {len(rows_to_delete)} duplicate rows from Excel file")
            return len(rows_to_delete), ws.max_row - 1
            
        except Exception as e:
            logger.error(f"Error cleaning Excel duplicates: {str(e)}")
            return 0, 0
    
    def get_file_info(self):
        """Get information about the Excel file."""
        if not os.path.exists(self.excel_file):
            return None
        
        try:
            file_stat = os.stat(self.excel_file)
            file_size = file_stat.st_size / 1024  # KB
            modified_time = datetime.fromtimestamp(file_stat.st_mtime)
            
            # Count unique records (not just max_row)
            wb = load_workbook(self.excel_file, read_only=True)
            ws = wb.active
            
            # Count unique applicant IDs to get accurate record count
            unique_ids = set()
            for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                if row[0]:
                    unique_ids.add(str(row[0]).strip())
            
            total_records = len(unique_ids)
            wb.close()
            
            return {
                'file_path': self.excel_file,
                'file_size': f"{file_size:.2f} KB",
                'total_records': total_records,
                'last_modified': modified_time.strftime("%Y-%m-%d %H:%M:%S")
            }
        except Exception as e:
            
            return None
    
    def import_from_excel(self, file_path, academic_year=None):
        """
        Import admission data from an uploaded Excel file.
        Expected columns: Applicant ID, Name, Programme, Status (optional)
        
        Args:
            file_path: Path to the uploaded Excel file
            academic_year: Academic year for the data (e.g., '2025/2026')
        
        Returns: (success, message, stats_dict)
        """
        try:
            from .models import KNUSTAdmission, get_current_academic_year
            
            # Use provided academic year or current
            if not academic_year:
                academic_year = get_current_academic_year()
            
            wb = load_workbook(file_path, read_only=True)
            ws = wb.active
            
            # Get headers from first row
            headers = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[1]]
            
            # Map column indices
            col_map = {}
            for idx, header in enumerate(headers):
                if 'applicant' in header or 'id' in header:
                    col_map['applicant_id'] = idx
                elif 'name' in header:
                    col_map['name'] = idx
                elif 'programme' in header or 'program' in header:
                    col_map['programme'] = idx
                elif 'status' in header or 'admission' in header or 'type' in header:
                    col_map['status'] = idx
                elif 'year' in header:
                    col_map['academic_year'] = idx
            
            # Validate required columns
            if 'applicant_id' not in col_map or 'name' not in col_map:
                wb.close()
                return False, "Excel file must have 'Applicant ID' and 'Name' columns", {}
            
            # Process rows
            added = 0
            updated = 0
            skipped = 0
            errors = []
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Extract values
                    applicant_id = str(row[col_map['applicant_id']]).strip() if row[col_map['applicant_id']] else None
                    name = str(row[col_map['name']]).strip() if row[col_map['name']] else None
                    
                    if not applicant_id or not name:
                        skipped += 1
                        continue
                    
                    # Get programme (optional)
                    programme = ''
                    if 'programme' in col_map and row[col_map['programme']]:
                        programme = str(row[col_map['programme']]).strip()
                    
                    # Get status (optional) - map to admission_status
                    admission_status = 'REGULAR'  # Default
                    if 'status' in col_map and row[col_map['status']]:
                        status_raw = str(row[col_map['status']]).strip().upper()
                        if 'FEE' in status_raw or 'PAYING' in status_raw:
                            admission_status = 'FEE_PAYING'
                        elif 'MATURE' in status_raw:
                            admission_status = 'MATURE'
                        elif 'INTERNATIONAL' in status_raw:
                            admission_status = 'INTERNATIONAL'
                        elif 'LESS' in status_raw or 'ENDOW' in status_raw:
                            admission_status = 'LESS_ENDOWED'
                        elif 'NMTC' in status_raw:
                            admission_status = 'NMTC'
                    
                    # Get academic year from row or use provided
                    row_academic_year = academic_year
                    if 'academic_year' in col_map and row[col_map['academic_year']]:
                        row_academic_year = str(row[col_map['academic_year']]).strip()
                    
                    # Create or update record
                    admission, created = KNUSTAdmission.objects.update_or_create(
                        applicant_id=applicant_id,
                        academic_year=row_academic_year,
                        defaults={
                            'name': name,
                            'programme': programme,
                            'admission_status': admission_status,
                            'source': 'UPLOAD',
                        }
                    )
                    
                    if created:
                        added += 1
                    else:
                        updated += 1
                        
                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
                    continue
            
            wb.close()
            
            stats = {
                'added': added,
                'updated': updated,
                'skipped': skipped,
                'errors': errors[:10],  # Limit errors shown
                'total_errors': len(errors),
                'academic_year': academic_year,
            }
            
            message = f"✅ Import complete: {added} added, {updated} updated, {skipped} skipped"
            if errors:
                message += f", {len(errors)} errors"
            
            return True, message, stats
            
        except Exception as e:
            error_msg = f"Error importing Excel file: {str(e)}"
            logger.error(error_msg)
            import traceback
            traceback.print_exc()
            return False, error_msg, {}
    
    def get_stats_by_academic_year(self):
        """Get admission statistics grouped by academic year."""
        from .models import KNUSTAdmission
        from django.db.models import Count
        
        stats = KNUSTAdmission.objects.values(
            'academic_year', 'programme_code', 'admission_status'
        ).annotate(count=Count('id')).order_by('-academic_year', 'programme_code')
        
        # Group by academic year
        year_stats = {}
        for stat in stats:
            year = stat['academic_year']
            if year not in year_stats:
                year_stats[year] = {
                    'total': 0,
                    'CS': 0,
                    'IT': 0,
                    'by_status': {}
                }
            year_stats[year]['total'] += stat['count']
            if stat['programme_code'] in ['CS', 'IT']:
                year_stats[year][stat['programme_code']] += stat['count']
            
            status = stat['admission_status']
            if status not in year_stats[year]['by_status']:
                year_stats[year]['by_status'][status] = 0
            year_stats[year]['by_status'][status] += stat['count']
        
        return year_stats
